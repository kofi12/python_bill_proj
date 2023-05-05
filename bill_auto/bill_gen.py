from openpyxl import Workbook
import os
import datetime

    
""" This module is used to generate a formatted excel spreadsheet of the billing data"""

# current month
today = datetime.date.today().strftime("%B")

# create a new workbook and 2 worksheets Billing and Attendance
wb = Workbook()
billing_sheet = wb.create_sheet("Billing")
attendance_sheet = wb.create_sheet("Attendance")

